"""
IOCF workbook parser.

Reads the "IOCF ALL BOARDS" Excel workbook and turns its very loosely
structured sheets (free-text cells, numbered lists, inline sub-headers)
into a JSON-friendly dict that the dashboard frontend can consume.

Nothing here is fabricated: every value returned is read straight out of
the workbook. Where a value can't be confidently identified it is simply
left out (never guessed).
"""
import re
import os
from datetime import datetime, date, time
import openpyxl

# ---------------------------------------------------------------------------
# Board sheet -> credits-sheet display name mapping
# ---------------------------------------------------------------------------
BOARD_SHEETS = {
    "Australia": "Australia",
    "Bangladesh": "Bangladesh",
    "Canada": "Canada",
    "England": "England",
    "India": "India",
    "Italy": "Italy",
    "Netherlands": "Netherlands",
    "Newzealand": "Newzealand",
    "Pakistan": "Pakistan",
    "Qatar": "Qatar",
    "Scotland": "Scotland",
    "South Africa": "SA",
    "Srilanka": "Srilanka",
    "UAE": "UAE",
    "West Indies": "WestIndies",
    "Zimbabwe": "Zimbabwe",
}

CREDITS_SHEET = "All Board Credits & Important U"

NUMBERED_RE = re.compile(r"^\s*\d+\.\s*(.+?)\s*$")
LEVEL_STADIUM_RE = re.compile(r"^level\s*\d+\s*stadiums?$", re.I)


def _clean(v):
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _num_or_none(v):
    if isinstance(v, (int, float)):
        return v
    return None


def _jsonify_scalar(v):
    """Make openpyxl cell values JSON-serialisable."""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, time):
        return v.strftime("%H:%M")
    return v


def _scalar_from_cell(cell):
    """Like _jsonify_scalar, but reconstructs "M-D" score text (e.g. "1-2")
    that Excel silently auto-converted into a date. Several rows in the
    "Series Result" column of the "Series Matches" sheet were typed as
    series scores (won-lost tallies) such as "1-2" or "3-2"; because that
    looks like a short date, Excel stored them as real dates formatted
    "m-d" with no year component. Reformatting those back to "M-D" avoids
    surfacing a misleading fabricated calendar date on the dashboard.
    """
    v = _clean(cell.value)
    if isinstance(v, datetime):
        nf = (cell.number_format or "").lower()
        if "y" not in nf and "m" in nf and "d" in nf:
            return f"{v.month}-{v.day}"
        return v.isoformat()
    if isinstance(v, (date,)):
        return v.isoformat()
    if isinstance(v, time):
        return v.strftime("%H:%M")
    return v


def col_letter(idx):
    return openpyxl.utils.get_column_letter(idx)


# ---------------------------------------------------------------------------
# Credits & rankings + stadium grid + tournament-update tables
# ---------------------------------------------------------------------------

def parse_credits_sheet(wb):
    ws = wb[CREDITS_SHEET]

    rankings = []
    r = 3
    while True:
        name = _clean(ws.cell(row=r, column=1).value)
        credits = _num_or_none(ws.cell(row=r, column=2).value)
        if name is None and credits is None:
            break
        if name:
            rankings.append({"board": name, "credits": credits})
        r += 1
    rankings.sort(key=lambda x: (x["credits"] or 0), reverse=True)
    for i, row in enumerate(rankings):
        row["rank"] = i + 1

    # Stadium tier grid, D1 header "Stadiums of all boards", D2:Q2 = board names
    stadium_grid = {}
    header_row = 2
    max_col = ws.max_column
    board_cols = {}
    for c in range(4, max_col + 1):
        name = _clean(ws.cell(row=header_row, column=c).value)
        if name:
            board_cols[c] = name
    for c, board in board_cols.items():
        entries = []
        r = 3
        while True:
            v = _clean(ws.cell(row=r, column=c).value)
            # stop once we've run past the grid block (row 16 is the last board row)
            if r > 16:
                break
            if v:
                entries.append(v)
            r += 1
        stadium_grid[board] = entries

    updates = _parse_updates_tables(ws)

    return {"rankings": rankings, "stadiumGrid": stadium_grid, "tournamentUpdates": updates}


TITLE_LIKE_RE = re.compile(r"^[A-Z0-9 '&\-().]+$")


def _looks_like_section_title(v):
    if not isinstance(v, str):
        return False
    v = v.strip()
    if len(v) < 4:
        return False
    if re.search(r"20\d\d", v):
        return True
    return bool(TITLE_LIKE_RE.match(v)) and v.upper() == v


def _parse_updates_tables(ws):
    """Parse the two side-by-side 'award / winner / board' recap tables
    (International Tournament Updates + Franchise League Updates) into a
    dict of tournament-name -> {champion, runnerUp, ...}."""
    tables = [
        # (start_row, name_col, award_col, winner_col, board_col)
        (24, 1, 2, 3, 4),   # A/B/C/D - international + emerging + lone warrior
        (24, 6, 7, 8, 10),  # F/G/H/J - franchise leagues
    ]
    sections = {}
    order = []
    for start_row, name_col, award_col, winner_col, board_col in tables:
        current = None
        for r in range(start_row, ws.max_row + 1):
            nameval = _clean(ws.cell(row=r, column=name_col).value)
            if _looks_like_section_title(nameval):
                current = nameval
                if current not in sections:
                    sections[current] = {}
                    order.append(current)
            award = _clean(ws.cell(row=r, column=award_col).value)
            winner = _clean(ws.cell(row=r, column=winner_col).value)
            board = _clean(ws.cell(row=r, column=board_col).value)
            if current and award:
                a = award.lower()
                if a.startswith("champion"):
                    sections[current]["champion"] = winner
                    sections[current]["champion_board"] = board
                elif a.startswith("runner"):
                    sections[current]["runnerUp"] = winner
                    sections[current]["runnerUp_board"] = board
    return {"order": order, "sections": sections}


# ---------------------------------------------------------------------------
# Individual board sheets
# ---------------------------------------------------------------------------

def _read_column_block(ws, col, start_row, end_row, blank_gap=1):
    """Read cell values straight down a column starting at start_row,
    stopping as soon as `blank_gap` consecutive empty cells are hit.

    Board sheets pack multiple unrelated tables into the same columns
    further down the sheet (e.g. a numbered player roster in column A
    rows 6-27, then - after a blank gap - an unrelated franchise-team
    roster reusing "1. Name" numbering starting at row 71). Stopping at
    the first blank run keeps each table isolated to its own block.
    """
    out = []
    blanks = 0
    for r in range(start_row, end_row + 1):
        v = _clean(ws.cell(row=r, column=col).value)
        if v is None:
            blanks += 1
            if blanks >= blank_gap:
                break
            continue
        blanks = 0
        out.append(v)
    return out


def _collect_numbered_list(ws, col, start_row, end_row):
    out = []
    for v in _read_column_block(ws, col, start_row, end_row):
        if isinstance(v, str):
            m = NUMBERED_RE.match(v)
            if m:
                out.append(m.group(1))
    return out


def _find_players_header_row(ws, max_row, default=5):
    """The "Players" column header is usually row 5, but a board sheet can
    grow an extra meta row above it (e.g. a "TITLE SPONSORS: ..." line) and
    shift everything below down by one. Scan for the actual "Players"
    (or "Players:") cell in column A instead of trusting a fixed row
    number, so the rest of the sheet's row offsets - credits, players,
    column headers - move with it.
    """
    for r in range(1, min(max_row, 15) + 1):
        v = _clean(ws.cell(row=r, column=1).value)
        if isinstance(v, str) and v.strip().rstrip(":").strip().lower() == "players":
            return r
    return default


def parse_board_sheet(wb, sheet_name, display_name, credits_fallback):
    ws = wb[sheet_name]
    max_col = ws.max_column
    max_row = ws.max_row

    title = _clean(ws.cell(row=1, column=1).value) or f"Board: {display_name}"
    meta_line = _clean(ws.cell(row=2, column=1).value) or ""

    chairman = None
    ceo = None
    m = re.search(r"chairman:\s*(.+?)(?:\s*\|\s*ceo:\s*(.+))?$", meta_line, re.I)
    if m:
        chairman = _clean(m.group(1))
        ceo = _clean(m.group(2))
    if chairman and "|" in chairman:
        # fallback split in case regex missed the CEO half
        parts = [p.strip() for p in chairman.split("|")]
        chairman = parts[0]

    header_row = _find_players_header_row(ws, max_row)

    credits = None
    for r in range(3, header_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, (int, float)):
            credits = v
            break
    if credits is None:
        credits = credits_fallback

    # Players: the numbered roster block in column A, right below the
    # "Players" header row. Stops at the first blank cell so it doesn't
    # bleed into unrelated franchise-team roster tables further down the
    # same column.
    players_start = header_row + 1
    players = _collect_numbered_list(ws, 1, players_start, max_row)
    if not players:
        players = _collect_numbered_list(ws, 1, players_start, min(max_row, 100))

    # Column headers live on the row found above (best-effort - not always
    # perfectly aligned)
    headers = {}
    for c in range(1, max_col + 1):
        v = _clean(ws.cell(row=header_row, column=c).value)
        if v:
            headers[c] = v

    stadium_col = None
    umpire_col = None
    for c, h in headers.items():
        low = h.lower()
        if "stadium" in low and stadium_col is None:
            stadium_col = c
        if "umpire" in low and umpire_col is None:
            umpire_col = c

    stadiums = []
    stadium_tier = None
    if stadium_col:
        for v in _read_column_block(ws, stadium_col, header_row + 1, max_row):
            if not isinstance(v, str):
                continue
            if LEVEL_STADIUM_RE.match(v):
                if stadium_tier is None:
                    stadium_tier = v
                continue
            if v.lower() == "none":
                continue
            stadiums.append(v)

    # Umpires: row 6 of the umpire column is a single "total credits
    # earned by every umpire on this board" figure. Row 7 is a header
    # ("Umpire" / "Details" / "Total Matches" / "Credits" / "Points")
    # across 5 columns starting at umpire_col. Rows 8+ are a numbered
    # list of umpire names in the base column; each name is followed by
    # one activity row per series/tournament/franchise-league they
    # officiated (category label + matches/credits/points for that
    # category), ending in a "Total Credits" row holding that umpire's
    # totals. A "Franchise League" label with no matches/credits/points
    # of its own is just a sub-header for the specific league rows right
    # below it (e.g. "BBL", "IPL") and isn't itself an activity. Two
    # boards (Italy, UAE) have no umpire column at all - umpires stays [].
    umpires = []
    umpire_credits = None
    if umpire_col:
        total_cell = ws.cell(row=header_row + 1, column=umpire_col).value
        if isinstance(total_cell, (int, float)):
            umpire_credits = total_cell
        current = None
        for r in range(header_row + 3, max_row + 1):
            name_cell = _clean(ws.cell(row=r, column=umpire_col).value)
            label = ws.cell(row=r, column=umpire_col + 1).value
            matches = _num_or_none(ws.cell(row=r, column=umpire_col + 2).value)
            credits_v = _num_or_none(ws.cell(row=r, column=umpire_col + 3).value)
            points_v = _num_or_none(ws.cell(row=r, column=umpire_col + 4).value)

            if isinstance(name_cell, str):
                m = NUMBERED_RE.match(name_cell)
                if m:
                    current = {
                        "name": m.group(1),
                        "totalCredits": None,
                        "totalPoints": None,
                        "activities": [],
                    }
                    umpires.append(current)

            if current is None or label is None:
                continue
            # "100 Ball League" gets auto-converted to the number 100 by
            # Excel since the sheet just writes the bare digits "100".
            if isinstance(label, (int, float)):
                label = "100 Balls" if float(label) == 100 else str(label)
            label = label.strip()
            if label.lower() == "total credits":
                current["totalCredits"] = credits_v
                current["totalPoints"] = points_v
            elif label == "Franchise League" and matches is None and credits_v is None:
                continue  # sub-header only - the specific leagues follow below
            else:
                current["activities"].append(
                    {"category": label, "matches": matches, "credits": credits_v, "points": points_v}
                )

    # Trophy cabinets + player-trade lists can appear as secondary inline
    # sub-headers anywhere in the sheet (not just row 5), e.g. a cell reading
    # "Trophy Cabinet ( Test )" partway down a column with a numbered list
    # directly below it. Scan every cell for these marker phrases.
    trophies = []
    transfers = []
    seen_trophy_markers = set()
    seen_transfer_markers = set()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = _clean(ws.cell(row=r, column=c).value)
            if not isinstance(v, str):
                continue
            low = v.lower()
            if "trophy cabinet" in low:
                key = (c, r)
                if key not in seen_trophy_markers:
                    seen_trophy_markers.add(key)
                    trophies.extend(_collect_numbered_list(ws, c, r + 1, max_row))
            elif "trade" in low and ":" in low:
                key = (c, r)
                if key not in seen_transfer_markers:
                    seen_transfer_markers.add(key)
                    transfers.extend(_collect_numbered_list(ws, c, r + 1, max_row))

    # de-dupe while preserving order
    def dedupe(seq):
        seen = set()
        out = []
        for x in seq:
            if x not in seen:
                seen.add(x)
                out.append(x)
        return out

    stadiums = dedupe(stadiums)
    trophies = dedupe(trophies)
    transfers = dedupe(transfers)

    return {
        "id": display_name.lower().replace(" ", "-"),
        "name": display_name,
        "title": title,
        "chairman": chairman,
        "ceo": ceo,
        "credits": credits,
        "playersCount": len(players),
        "players": players,
        "stadiumTier": stadium_tier,
        "stadiums": stadiums,
        "trophies": trophies,
        "trophiesCount": len(trophies),
        "umpires": umpires,
        "umpiresCount": len(umpires),
        "umpireCredits": umpire_credits,
        "transfers": transfers,
    }


def parse_dismantled_sheet(wb, sheet_name):
    """Archived/dismantled boards - kept separate from the 14 active boards."""
    ws = wb[sheet_name]
    title = _clean(ws.cell(row=1, column=1).value) or sheet_name
    meta_line = _clean(ws.cell(row=2, column=1).value) or ""
    chairman = None
    m = re.search(r"chairman:\s*(.+)$", meta_line, re.I)
    if m:
        chairman = _clean(m.group(1))
    credits_line = _clean(ws.cell(row=3, column=1).value) or ""
    credits = None
    m2 = re.search(r"([\d,]+)", credits_line)
    if m2:
        try:
            credits = int(m2.group(1).replace(",", ""))
        except ValueError:
            pass
    players = _collect_numbered_list(ws, 1, 5, ws.max_row)
    return {
        "name": title.replace("Board:", "").strip(),
        "chairman": chairman,
        "credits": credits,
        "playersCount": len(players),
    }


def get_boards(wb):
    credits_data = parse_credits_sheet(wb)
    credits_by_name = {r["board"]: r["credits"] for r in credits_data["rankings"]}
    rank_by_name = {r["board"]: r["rank"] for r in credits_data["rankings"]}

    boards = []
    for display_name, sheet_name in BOARD_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        board = parse_board_sheet(
            wb, sheet_name, display_name, credits_by_name.get(display_name)
        )
        board["ranking"] = rank_by_name.get(display_name)
        boards.append(board)

    boards.sort(key=lambda b: (b["ranking"] or 999))

    dismantled = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower().startswith("dismantled"):
            dismantled.append(parse_dismantled_sheet(wb, sheet_name))

    return boards, dismantled, credits_data


# ---------------------------------------------------------------------------
# T20 World Cup 2026 bracket
# ---------------------------------------------------------------------------

def get_t20_world_cup(wb):
    name = "T20 World Cup 2026"
    if name not in wb.sheetnames:
        return None
    ws = wb[name]
    max_row = ws.max_row

    title = _clean(ws.cell(row=1, column=1).value)

    groups = {}
    r = 2
    header_row = None
    for rr in (2,):
        vals = [_clean(ws.cell(row=rr, column=c).value) for c in range(1, 4)]
        if any(vals):
            header_row = rr
    if header_row:
        group_names = {}
        for c in range(1, 4):
            v = _clean(ws.cell(row=header_row, column=c).value)
            if v:
                group_names[c] = v
        for c, gname in group_names.items():
            teams = []
            rr = header_row + 1
            while True:
                v = _clean(ws.cell(row=rr, column=c).value)
                if v is None:
                    break
                m = NUMBERED_RE.match(v) if isinstance(v, str) else None
                teams.append(m.group(1) if m else v)
                rr += 1
                if rr - header_row > 12:
                    break
            groups[gname] = teams

    def read_match_block(start_row):
        """Reads a 'Schedule/Winner/Host/MoM/Best Bat/Best Bowl' style block."""
        header_r = start_row
        headers = {}
        for c in range(1, 7):
            v = _clean(ws.cell(row=header_r, column=c).value)
            if v:
                headers[c] = v
        matches = []
        rr = header_r + 1
        while rr <= max_row:
            a = _clean(ws.cell(row=rr, column=1).value)
            if a is None:
                break
            if not isinstance(a, str) or "vs" not in a.lower():
                break
            row = {}
            for c, h in headers.items():
                v = _jsonify_scalar(_clean(ws.cell(row=rr, column=c).value))
                if v is not None:
                    row[h.strip().rstrip(":").strip()] = v
            matches.append(row)
            rr += 1
        return matches, rr

    stages = {}
    r = 1
    stage_markers = [
        "Group A Matches",
        "Group B Matches",
        "Super 6",
        "Semifinal",
        "Final",
    ]
    while r <= max_row:
        a = _clean(ws.cell(row=r, column=1).value)
        if a in stage_markers:
            # header row for the block is usually the very next row
            matches, next_r = read_match_block(r + 1)
            stages[a] = matches
            r = next_r
            continue
        r += 1

    # Awards table
    awards = []
    for r in range(1, max_row + 1):
        a = _clean(ws.cell(row=r, column=1).value)
        if a == "Awards":
            hr = r + 1
            rr = hr + 1
            while rr <= max_row:
                award = _clean(ws.cell(row=rr, column=1).value)
                if not award:
                    break
                awards.append(
                    {
                        "award": award,
                        "winner": _clean(ws.cell(row=rr, column=2).value),
                        "board": _clean(ws.cell(row=rr, column=3).value),
                        "credits": _num_or_none(ws.cell(row=rr, column=4).value),
                    }
                )
                rr += 1
            break

    champion = None
    runner_up = None
    if "Final" in stages and stages["Final"]:
        final = stages["Final"][0]
        champion = final.get("Winner")
        teams = [t.strip() for t in re.split(r"\bvs\b", final.get("Schedule", ""), flags=re.I)]
        if champion and len(teams) == 2:
            runner_up = teams[0] if teams[1] == champion else teams[1]

    total_matches = sum(len(v) for v in stages.values())

    return {
        "title": title,
        "groups": groups,
        "stages": stages,
        "awards": awards,
        "champion": champion,
        "runnerUp": runner_up,
        "totalMatches": total_matches,
        "status": "Completed" if champion else "Ongoing",
    }


# ---------------------------------------------------------------------------
# Series Matches & WTC Matches (bilateral fixtures)
# ---------------------------------------------------------------------------

KNOWN_TITLES = {
    "Completed Series",
    "Upcoming Series",
    "World Test Championship Completed Matches",
    "Upcoming IOCF World Test Championship Matches",
    "Franchise Leagues",
    "Major Tournaments",
    "WTC PENDING SCHEDULE",
}

MONTHS = {
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
}


def _find_titles(ws):
    found = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = _clean(ws.cell(row=r, column=c).value)
            if isinstance(v, str) and v.strip() in KNOWN_TITLES:
                found.append((r, c, v.strip()))
    return found


def _read_table_block(ws, title_row, title_col, next_title_row):
    header_row = title_row + 1
    headers = {}
    c = title_col
    while True:
        v = _clean(ws.cell(row=header_row, column=c).value)
        if v is None:
            if c > title_col and c - title_col > 20:
                break
            if v is None and c > title_col + 12:
                break
        if v:
            headers[c] = v.strip()
            c += 1
        else:
            if headers:
                break
            c += 1
            if c - title_col > 15:
                break
    if not headers:
        return []

    end_row = next_title_row - 1 if next_title_row else ws.max_row
    rows = []
    blank_streak = 0
    for r in range(header_row + 1, end_row + 1):
        cells = {h: _scalar_from_cell(ws.cell(row=r, column=c)) for c, h in headers.items()}
        non_null = {k: v for k, v in cells.items() if v is not None}
        if not non_null:
            blank_streak += 1
            if blank_streak > 3:
                break
            continue
        blank_streak = 0
        # a lone month-name divider row - skip it, not real data
        first_val = list(non_null.values())[0]
        if len(non_null) == 1 and isinstance(first_val, str) and first_val.strip().lower() in MONTHS:
            continue
        rows.append(non_null)
    return rows


def _parse_fixture_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    titles = _find_titles(ws)
    result = {}
    for i, (row, col, title) in enumerate(titles):
        next_row = None
        for j in range(i + 1, len(titles)):
            if titles[j][0] > row:
                next_row = titles[j][0]
                break
        data = _read_table_block(ws, row, col, next_row)
        key = re.sub(r"[^a-zA-Z0-9]+", " ", title).strip().title().replace(" ", "")
        key = key[0].lower() + key[1:]
        if key in result:
            result[key] = result[key] + data
        else:
            result[key] = data
    return result


def get_fixtures(wb):
    series = _parse_fixture_sheet(wb, "Series Matches")
    tests = _parse_fixture_sheet(wb, "WTC Matches")
    return {"series": series, "tests": tests}


def _levenshtein(a, b):
    if a == b:
        return 0
    la, lb = len(a), len(b)
    if la == 0:
        return lb
    if lb == 0:
        return la
    prev = list(range(lb + 1))
    for i, ca in enumerate(a, 1):
        cur = [i] + [0] * lb
        for j, cb in enumerate(b, 1):
            cost = 0 if ca == cb else 1
            cur[j] = min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + cost)
        prev = cur
    return prev[lb]


# Afghanistan was dissolved as an IOCF board and replaced by Qatar - the
# same players carried over, so its historical fixture rows ("Afghanistan"
# as Hosting Board/Opponents/Winners) resolve to Qatar rather than being
# dropped as an unregistered board. A future, newly-formed Afghanistan
# board would be a distinct entity recorded under its own sheet, so this
# alias only ever matches the literal historical spelling here.
_BOARD_NAME_ALIASES = {
    "afghanistan": "Qatar",
}


def _resolve_board_name(name):
    """Resolve a free-text board name from a fixture sheet to one of the 14
    known boards - exact match first, then the historical alias table
    above, then a case-insensitive Levenshtein distance of 1 (catches typos
    like "Netherland"). Multi-team strings ("A, B & C") and boards that were
    never registered resolve to None rather than being guessed at.
    """
    if not isinstance(name, str):
        return None
    cleaned = name.strip()
    if not cleaned:
        return None
    for board in BOARD_SHEETS:
        if cleaned.lower() == board.lower():
            return board
    alias = _BOARD_NAME_ALIASES.get(cleaned.lower())
    if alias:
        return alias
    match = None
    for board in BOARD_SHEETS:
        if _levenshtein(cleaned.lower(), board.lower()) <= 1:
            if match is not None:
                return None
            match = board
    return match


_SCORE_RE = re.compile(r"^(\d+)-(\d+)$")


def _series_match_count(result):
    """How many individual matches a completed series' "Series Result" cell
    represents - not 1 per series. Most results are a clean "N-M" score
    (matches played = N + M), but a handful of sheets instead spell out
    "1 Match only" or list each match on its own line (a mini round-robin +
    final squeezed into one series row) - both are counted from what's
    actually written rather than assumed to be a single match."""
    if not isinstance(result, str):
        return 0
    text = result.strip()
    if not text:
        return 0
    m = _SCORE_RE.match(text.replace(" ", ""))
    if m:
        return int(m.group(1)) + int(m.group(2))
    m = re.match(r"^(\d+)\s*match", text, re.I)
    if m:
        return int(m.group(1))
    lines = [ln for ln in text.split("\n") if ln.strip()]
    if len(lines) > 1:
        vs_lines = [ln for ln in lines if " vs " in ln.lower()]
        if vs_lines:
            return len(vs_lines)
    return 0


def get_head_to_head(wb):
    """Every clean two-team meeting (series or Test) between two of the 14
    known boards, for the board comparator / rivalry history page. Rows
    naming a multi-team series ("A, B & C") or a host/opponent that is not
    one of the 14 known boards are dropped rather than guessed at.
    """
    fixtures = get_fixtures(wb)
    meetings = []

    def _clean_pair(host_raw, opponents_raw):
        if not isinstance(host_raw, str) or "," in host_raw or "&" in host_raw:
            return None, None
        if not isinstance(opponents_raw, str) or "," in opponents_raw or "&" in opponents_raw:
            return None, None
        host = _resolve_board_name(host_raw)
        opponent = _resolve_board_name(opponents_raw)
        if not host or not opponent or host == opponent:
            return None, None
        return host, opponent

    for row in fixtures["series"].get("completedSeries", []):
        host, opponent = _clean_pair(row.get("Hosting Board"), row.get("Opponents"))
        if not host:
            continue
        host_score = opponent_score = None
        result = row.get("Series Result")
        if isinstance(result, str):
            m = _SCORE_RE.match(result.replace(" ", ""))
            if m:
                host_score, opponent_score = int(m.group(1)), int(m.group(2))
        meetings.append(
            {
                "format": "Series",
                "name": row.get("Series Name"),
                "dates": row.get("Dates"),
                "hostBoard": host,
                "opponentBoard": opponent,
                "winner": _resolve_board_name(row.get("Winners")),
                "hostScore": host_score,
                "opponentScore": opponent_score,
            }
        )

    for row in fixtures["tests"].get("worldTestChampionshipCompletedMatches", []):
        host, opponent = _clean_pair(row.get("Hosting Board"), row.get("Opponents"))
        if not host:
            continue
        meetings.append(
            {
                "format": "Test",
                "name": row.get("Test Name"),
                "dates": row.get("Dates"),
                "hostBoard": host,
                "opponentBoard": opponent,
                "winner": _resolve_board_name(row.get("Winners")),
                "hostScore": None,
                "opponentScore": None,
            }
        )

    return meetings


def get_records(wb):
    """Biggest wins, whitewashes and most-decorated-player tallies pulled
    straight from the completed Series/Test sheets - only rows with a clean
    two-team, clean-numeric result count toward a margin, nothing guessed.
    """
    fixtures = get_fixtures(wb)
    biggest_wins = []
    whitewashes = []

    for row in fixtures["series"].get("completedSeries", []):
        opponents_raw = row.get("Opponents")
        host_raw = row.get("Hosting Board")
        if not isinstance(opponents_raw, str) or "," in opponents_raw or "&" in opponents_raw:
            continue
        if not isinstance(host_raw, str) or "," in host_raw or "&" in host_raw:
            continue
        result = row.get("Series Result")
        if not isinstance(result, str):
            continue
        m = _SCORE_RE.match(result.replace(" ", ""))
        if not m:
            continue
        host_score, opponent_score = int(m.group(1)), int(m.group(2))
        entry = {
            "seriesName": row.get("Series Name"),
            "hostBoard": host_raw,
            "opponentBoard": opponents_raw,
            "winner": row.get("Winners"),
            "loser": row.get("Runners"),
            "result": result,
            "margin": abs(host_score - opponent_score),
            "dates": row.get("Dates"),
        }
        biggest_wins.append(entry)
        if host_score == 0 or opponent_score == 0:
            whitewashes.append(entry)

    biggest_wins.sort(key=lambda e: e["margin"], reverse=True)
    whitewashes.sort(key=lambda e: e["margin"], reverse=True)

    # Closest matches: the same completed-series pool, sorted the other way
    # and excluding whitewashes (a 0-margin sweep isn't a "nail-biter").
    closest_matches = sorted(
        (e for e in biggest_wins if e["margin"] > 0),
        key=lambda e: e["margin"],
    )

    def _tally(field):
        counts = {}
        order = []
        for row in fixtures["tests"].get("worldTestChampionshipCompletedMatches", []):
            v = row.get(field)
            if not isinstance(v, str) or not v.strip():
                continue
            v = v.strip()
            if v not in counts:
                counts[v] = 0
                order.append(v)
            counts[v] += 1
        ranked = [{"player": name, "count": counts[name]} for name in order]
        ranked.sort(key=lambda e: e["count"], reverse=True)
        return ranked

    return {
        "biggestWins": biggest_wins[:10],
        "whitewashes": whitewashes[:10],
        "closestMatches": closest_matches[:10],
        "manOfTheMatch": _tally("Man of the Match")[:10],
        "bestBatsman": _tally("Best Batsman")[:10],
        "bestBowler": _tally("Best Bowler")[:10],
    }


# ---------------------------------------------------------------------------
# Franchise league squad sheets
# ---------------------------------------------------------------------------

FRANCHISE_SHEETS = [
    "100 2026",
    "PSL 2026",
    "BBL 2026",
    "KCL 2026",
    "BPL 2026",
    "IPL 2026",
    "CPL 2026",
    "SPL 2026",
]

FRANCHISE_DISPLAY_NAMES = {
    "100 2026": "The Hundred 2026",
    "PSL 2026": "Pakistan Super League 2026",
    "BBL 2026": "Big Bash League 2026",
    "KCL 2026": "Kiwi Crown League 2026",
    "BPL 2026": "Bangladesh Premier League 2026",
    "IPL 2026": "Indian Premier League 2026",
    "CPL 2026": "Caribbean Premier League 2026",
    "SPL 2026": "Scottish Premier League 2026",
}

FRANCHISE_ALL_TEAMS_RE = re.compile(r"all\s*teams\s*$", re.I)
FRANCHISE_AWARDS_TITLE_RE = re.compile(r"\bawards\b", re.I)
FRANCHISE_CANCELLED_RE = re.compile(r"cancelled", re.I)

# Player-roster lines in a franchise team's own column look like
# "1. Shibasis (c) - 8000", "4. Manish Indoria - Aus - 2500", or (CPL only)
# "1. Shibil (WI-C)" with no credits at all for the top few slots. This one
# regex + a bit of post-processing below covers every variant actually
# found in the workbook rather than assuming one clean format.
PLAYER_LINE_RE = re.compile(r"^\s*\d+\.\s*(.+?)\s*$")
TRAILING_CREDITS_RE = re.compile(r"(\d[\d,]*)\s*([kK])?\s*$")
REPLACED_NOTE_RE = re.compile(r"\(\s*replaced[^)]*\)\s*$", re.I)
ROLE_TAG_RE = re.compile(r"\((c|vc|m|dc|ds)\)\s*$", re.I)
# Role can appear on either side of the hyphen - "(WI-C)" (board-role) or
# "(DS-Ind)" (role-board) both occur in the workbook, so both groups are
# checked against ROLE_ABBRS rather than assuming a fixed position.
ROLE_PAIR_RE = re.compile(r"\(([a-z]{1,5})-([a-z]{1,5})\)\s*$", re.I)
ROLE_ABBRS = {"c", "vc", "m", "dc", "ds"}
# Confirmed against the source workbook: (c) = Captain, (m) = Marquee (a
# pre-selected marquee-category signing, not a mentor), (ds) = Direct
# Signing (signed directly by the franchise outside the main auction).
ROLE_NAMES = {"c": "Captain", "vc": "Vice-Captain", "m": "Marquee", "dc": "Designated Coach", "ds": "Direct Signing"}


def _parse_franchise_player_line(raw):
    """Parses one numbered roster line from a "Franchise League ... All
    TEAMS" block into {name, credits, role, note}. Deliberately tolerant:
    the source workbook mixes several hand-typed formats (see the sampled
    variants above PLAYER_LINE_RE) rather than one consistent template, so
    this extracts what it confidently can and leaves the rest as None
    instead of guessing.
    """
    m = PLAYER_LINE_RE.match(raw)
    text = m.group(1) if m else raw.strip()

    note = None
    note_m = REPLACED_NOTE_RE.search(text)
    if note_m:
        note = _clean(note_m.group(0))
        text = text[: note_m.start()].rstrip()

    # Credits are stripped before the role tag: most lines put the role
    # tag right after the name and the credits at the very end
    # ("Shibasis (c) - 8000"), so a role check anchored to end-of-string
    # would miss it if it ran before the trailing credits were removed.
    credits = None
    credits_m = TRAILING_CREDITS_RE.search(text)
    if credits_m:
        num = int(credits_m.group(1).replace(",", ""))
        if credits_m.group(2):
            num *= 1000
        credits = num
        text = text[: credits_m.start()].rstrip(" -").strip()

    role = None
    pair_m = ROLE_PAIR_RE.search(text)
    if pair_m:
        a, b = pair_m.group(1), pair_m.group(2)
        a_l, b_l = a.lower(), b.lower()
        remainder = None
        if a_l in ROLE_ABBRS and b_l not in ROLE_ABBRS:
            role, remainder = ROLE_NAMES[a_l], b
        elif b_l in ROLE_ABBRS and a_l not in ROLE_ABBRS:
            role, remainder = ROLE_NAMES[b_l], a
        elif a_l in ROLE_ABBRS:
            role = ROLE_NAMES[a_l]
        elif b_l in ROLE_ABBRS:
            role = ROLE_NAMES[b_l]
        if role is not None:
            text = text[: pair_m.start()].rstrip()
            # The other half of the pair is usually the board the pick
            # actually represents ("(M-Aus)" = Marquee + Australia) - the
            # frontend already parses trailing "(XXX)" board tags off other
            # roster lines that never had a role attached, so keeping this
            # one lets it resolve the same way instead of losing the board
            # info here and never being able to recover it downstream.
            if remainder:
                text = f"{text}({remainder})"
    if role is None:
        role_m = ROLE_TAG_RE.search(text)
        if role_m:
            role = ROLE_NAMES.get(role_m.group(1).lower())
            text = text[: role_m.start()].rstrip()

    name = _clean(text) or _clean(raw)
    return {"name": name, "credits": credits, "role": role, "note": note}


def _find_row_with_text(ws, pattern, start_row=1, end_row=None, column=1):
    """Returns the first row number (>= start_row) whose given column value
    matches `pattern` (a compiled regex tested against the cleaned string
    value), or None."""
    end_row = end_row or ws.max_row
    for r in range(start_row, end_row + 1):
        v = _clean(ws.cell(row=r, column=column).value)
        if isinstance(v, str) and pattern.search(v):
            return r
    return None


def _read_franchise_teams(ws, header_row, end_row):
    """Reads the "Franchise League X All TEAMS" block starting at
    `header_row` (the row holding each team's name across columns) through
    `end_row` (exclusive) - one column per team, one numbered player-line
    per row below the header. The very first non-numbered row right below
    the header (if present) is the "Remaining Purse - N" line some sheets
    include per team.
    """
    teams = {}
    for c in range(1, ws.max_column + 1):
        team_name = _clean(ws.cell(row=header_row, column=c).value)
        if not team_name:
            continue

        remaining_purse = None
        roster = []
        for r in range(header_row + 1, end_row):
            raw = _clean(ws.cell(row=r, column=c).value)
            if raw is None:
                continue
            if isinstance(raw, str) and raw.lower().startswith("remaining purse"):
                purse_m = TRAILING_CREDITS_RE.search(raw)
                if purse_m:
                    remaining_purse = int(purse_m.group(1).replace(",", ""))
                continue
            if isinstance(raw, str) and PLAYER_LINE_RE.match(raw):
                roster.append(_parse_franchise_player_line(raw))

        if roster:
            teams[team_name] = {
                "players": roster,
                "remainingPurse": remaining_purse,
            }
    return teams


def _read_franchise_matches(ws, header_row, end_row):
    """Reads a "Schedule / Date / [Venue] / Winner / Umpire / Man of the
    Match / Best Batsman / [Best Bowler]" table - column layout genuinely
    varies between franchise sheets (some have a Venue column, some have a
    separate Best Bowler column, some don't), so headers are read directly
    from `header_row` rather than assumed."""
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = _clean(ws.cell(row=header_row, column=c).value)
        if h:
            headers[c] = h.strip().rstrip(":").strip()

    matches = []
    for r in range(header_row + 1, end_row):
        schedule = _clean(ws.cell(row=r, column=1).value)
        if not schedule:
            continue
        row = {}
        for c, h in headers.items():
            v = _jsonify_scalar(_clean(ws.cell(row=r, column=c).value))
            if v is not None:
                row[h] = v
        if row:
            matches.append(row)
    return matches


def _read_franchise_awards(ws, header_row, end_row):
    """Reads the "Award Name / Award Winner / Franchise Team / National
    Board / Award Credits" table. The final couple of rows are usually
    "Runners up" / "Champions" (no franchise/board columns, just a winner
    name + credits) - those are captured the same way as any other award
    row rather than special-cased, and pulled back out by name below for
    champion/runnerUp."""
    awards = []
    for r in range(header_row + 1, end_row):
        award_name = _clean(ws.cell(row=r, column=1).value)
        if not award_name:
            continue
        awards.append(
            {
                "award": award_name,
                "winner": _clean(ws.cell(row=r, column=2).value),
                "team": _clean(ws.cell(row=r, column=3).value),
                "board": _clean(ws.cell(row=r, column=4).value),
                "credits": _num_or_none(ws.cell(row=r, column=5).value),
            }
        )
    return awards


def get_franchise_leagues(wb):
    leagues = []
    for sheet_name in FRANCHISE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        max_row = ws.max_row
        title = _clean(ws.cell(row=1, column=1).value) or sheet_name
        display_name = FRANCHISE_DISPLAY_NAMES.get(sheet_name, sheet_name)

        # National-board player registrations: every "<Board name>" header
        # row followed by a numbered list, before the "All TEAMS" section.
        # A sheet can have more than one such header row side by side in
        # time (e.g. The Hundred lists 6 boards, then 6 more starting a
        # further 14 rows down) - so this scans every row rather than
        # assuming a single fixed header row.
        all_teams_row = _find_row_with_text(ws, FRANCHISE_ALL_TEAMS_RE, start_row=2)
        registration_end = all_teams_row or max_row

        boards = {}
        for r in range(2, registration_end):
            next_row_first_cell = _clean(ws.cell(row=r + 1, column=1).value)
            if not (isinstance(next_row_first_cell, str) and next_row_first_cell.strip().startswith("1.")):
                continue
            for c in range(1, ws.max_column + 1):
                board = _clean(ws.cell(row=r, column=c).value)
                if not board or (isinstance(board, str) and board[0].isdigit()):
                    continue
                players = [p["name"] for p in (
                    _parse_franchise_player_line(v)
                    for v in _read_column_block(ws, c, r + 1, registration_end - 1)
                    if isinstance(v, str) and PLAYER_LINE_RE.match(v)
                )]
                if players:
                    boards.setdefault(board, players)

        # Franchise team rosters, match schedule, awards - each section
        # runs from its own title row to the next title row (or end of
        # sheet). Any section whose title isn't found is simply empty
        # rather than raising - several leagues are missing one or more
        # sections because the season itself is incomplete/cancelled in
        # the source workbook (see `status` below).
        teams = {}
        matches = []
        awards = []
        if all_teams_row:
            match_title_row = _find_row_with_text(ws, re.compile(r"all match updates", re.I), start_row=all_teams_row)
            awards_title_row = _find_row_with_text(ws, FRANCHISE_AWARDS_TITLE_RE, start_row=all_teams_row)
            cancelled_row = _find_row_with_text(ws, FRANCHISE_CANCELLED_RE, start_row=all_teams_row)

            teams_end = match_title_row or awards_title_row or cancelled_row or (max_row + 1)
            teams = _read_franchise_teams(ws, all_teams_row + 1, teams_end)

            if match_title_row:
                matches_end = awards_title_row or cancelled_row or (max_row + 1)
                matches = _read_franchise_matches(ws, match_title_row + 1, matches_end)

            if awards_title_row:
                awards = _read_franchise_awards(ws, awards_title_row + 1, max_row + 1)

        # Champion/runner-up read straight out of this sheet's own awards
        # table (the last two rows are always "Champions"/"Runners up" when
        # present) rather than cross-referencing a separate credits-sheet
        # summary table - self-contained per league and doesn't depend on
        # display-name spelling matching across two sheets.
        champion = None
        runner_up = None
        for a in awards:
            name = (a["award"] or "").strip().lower()
            if name == "champions":
                champion = a["winner"]
            elif name in ("runners up", "runner up", "runners-up"):
                runner_up = a["winner"]

        is_cancelled = _find_row_with_text(ws, FRANCHISE_CANCELLED_RE, start_row=2) is not None
        if is_cancelled:
            status = "Cancelled"
        elif champion:
            status = "Completed"
        elif matches:
            status = "Ongoing"
        else:
            status = "Upcoming"

        leagues.append(
            {
                "id": sheet_name.lower().replace(" ", "-"),
                "name": display_name,
                "title": title,
                "season": "2026",
                "boards": boards,
                "teams": teams,
                "teamCount": len(teams) or len(boards),
                "matches": matches,
                "totalMatches": len(matches),
                "awards": awards,
                "champion": champion,
                "runnerUp": runner_up,
                "status": status,
            }
        )
    return leagues
    return leagues


# ---------------------------------------------------------------------------
# Hall of Fame cards
# ---------------------------------------------------------------------------

def get_hall_of_fame(wb):
    name = "IOCF Cards"
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    cards = []
    current = None
    for r in range(1, ws.max_row + 1):
        a = _clean(ws.cell(row=r, column=1).value)
        if isinstance(a, str) and a.lower().startswith("card name"):
            if current:
                cards.append(current)
            current = {
                "name": a.split(":", 1)[1].strip() if ":" in a else a,
                "subtitle": None,
                "players": [],
            }
            continue
        if current and current["subtitle"] is None and isinstance(a, str) and not a.lower().startswith("player"):
            current["subtitle"] = a
            continue
        if current and isinstance(a, str) and a.lower() == "player name":
            continue
        if current and a and a not in ("Player Name",):
            country = _clean(ws.cell(row=r, column=2).value)
            award = _clean(ws.cell(row=r, column=3).value)
            achievement = _clean(ws.cell(row=r, column=4).value)
            if country or award:
                current["players"].append(
                    {
                        "name": a,
                        "country": country,
                        "award": award,
                        "achievement": achievement,
                    }
                )
    if current:
        cards.append(current)
    return cards


# ---------------------------------------------------------------------------
# Board Rankings — a genuine points-based performance ranking, one table per
# format (T10/T20/Club/ODI/100 Balls/Test) plus an Overall table, separate
# from the credits balance that powers the Credits Ranking page.
# ---------------------------------------------------------------------------

def _num_or_none_rankings(v):
    """Like _num_or_none, but also treats Excel error strings
    ("#DIV/0!" - a board with 0 matches played divides by zero computing
    its winning rate/scores) as missing rather than fabricating a 0."""
    if isinstance(v, (int, float)):
        return v
    return None


def _parse_ranking_table(ws, header_row, max_row):
    rows = []
    r = header_row + 1
    while r <= max_row:
        board = _clean(ws.cell(row=r, column=1).value)
        if board is None:
            break
        details = _clean(ws.cell(row=r, column=4).value)
        rows.append(
            {
                "board": board,
                "matchesPlayed": _num_or_none_rankings(ws.cell(row=r, column=2).value),
                "matchesWon": _num_or_none_rankings(ws.cell(row=r, column=3).value),
                "honors": None if details in (None, "NA") else details,
                "winningRate": _num_or_none_rankings(ws.cell(row=r, column=5).value),
                "leagueScore": _num_or_none_rankings(ws.cell(row=r, column=6).value),
                "tournamentScore": _num_or_none_rankings(ws.cell(row=r, column=7).value),
                "points": _num_or_none_rankings(ws.cell(row=r, column=8).value),
            }
        )
        r += 1
    # Already sorted by Points in the sheet, but assign explicit ranks
    # rather than relying on row order alone.
    rows.sort(key=lambda x: (x["points"] if x["points"] is not None else -1), reverse=True)
    for i, row in enumerate(rows):
        row["rank"] = i + 1
    return rows, r


def get_board_rankings(wb):
    name = "Board Rankings"
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    max_row = ws.max_row
    tables = []
    r = 1
    while r <= max_row:
        title = _clean(ws.cell(row=r, column=1).value)
        if isinstance(title, str) and title.lower().endswith("ranking"):
            header = _clean(ws.cell(row=r + 1, column=1).value)
            if isinstance(header, str) and header.strip().lower() == "board":
                rows, next_r = _parse_ranking_table(ws, r + 1, max_row)
                tables.append(
                    {
                        "id": title.lower().replace(" ", "-"),
                        "name": title,
                        "table": rows,
                    }
                )
                r = next_r
                continue
        r += 1
    return tables


def get_umpire_rankings(boards):
    """Global points-based umpire leaderboard, built from the per-board
    umpire data every board sheet now carries. A handful of umpires
    officiate for more than one board and appear on both boards' sheets
    under the exact same spelling - those are merged into a single
    ranked entry (summed credits/points, activities tagged with which
    board they were earned for) rather than counted twice. Merging is by
    exact name match only; anything not an exact match is kept separate
    rather than guessed at.
    """
    by_name = {}
    order = []
    for board in boards:
        for umpire in board.get("umpires", []):
            name = umpire["name"]
            entry = by_name.get(name)
            if entry is None:
                entry = {
                    "name": name,
                    "boards": [],
                    "totalCredits": None,
                    "totalPoints": None,
                    "activities": [],
                }
                by_name[name] = entry
                order.append(name)
            entry["boards"].append(board["name"])
            if umpire["totalCredits"] is not None:
                entry["totalCredits"] = (entry["totalCredits"] or 0) + umpire["totalCredits"]
            if umpire["totalPoints"] is not None:
                entry["totalPoints"] = (entry["totalPoints"] or 0) + umpire["totalPoints"]
            for activity in umpire["activities"]:
                entry["activities"].append({**activity, "board": board["name"]})

    rankings = [by_name[name] for name in order]
    rankings.sort(key=lambda x: (x["totalPoints"] if x["totalPoints"] is not None else -1), reverse=True)
    for i, entry in enumerate(rankings):
        entry["rank"] = i + 1
    return rankings


# ---------------------------------------------------------------------------
# Continental cups + emerging talent + lone warrior (participant lists)
# ---------------------------------------------------------------------------

CONTINENTAL_CUPS = {
    "IOCF ASIA CUP": "IOCF Asia Cup",
    "IOCF EURO CUP": "IOCF Euro Cup",
    "IOCF OCEANIA CUP": "IOCF Oceania Cup",
}

CONTINENTAL_CUP_SCHEDULE_RE = re.compile(r"schedule and all match updates", re.I)


def get_continental_cups(wb):
    cups = []
    for sheet_name, display in CONTINENTAL_CUPS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        teams = []
        for r in range(1, ws.max_row + 1):
            v = _clean(ws.cell(row=r, column=1).value)
            if isinstance(v, str):
                m = NUMBERED_RE.match(v)
                if m:
                    teams.append(m.group(1))

        matches = []
        schedule_row = _find_row_with_text(ws, CONTINENTAL_CUP_SCHEDULE_RE)
        if schedule_row:
            matches = _read_franchise_matches(ws, schedule_row + 1, ws.max_row + 1)

        completed = [m for m in matches if m.get("Winner")]
        if matches and len(completed) == len(matches):
            status = "Completed"
        elif completed:
            status = "Ongoing"
        else:
            status = "Upcoming"

        cups.append(
            {
                "id": sheet_name.lower().replace(" ", "-"),
                "name": display,
                "season": "2026",
                "teams": teams,
                "matches": matches,
                "totalMatches": len(matches),
                "status": status,
                "champion": None,
            }
        )
    return cups


def get_emerging_talent_league(wb, tournament_updates):
    name = "Emerging Talent League"
    info = tournament_updates.get("sections", {}).get("EMERGING TALENTS LEAGUE 2026", {})
    squads = {}
    matches = []
    if name in wb.sheetnames:
        ws = wb[name]
        for c in range(1, 5):
            board = _clean(ws.cell(row=2, column=c).value)
            if board:
                squads[board] = _collect_numbered_list(ws, c, 3, ws.max_row)
        for r in range(3, ws.max_row + 1):
            host = _clean(ws.cell(row=r, column=6).value)
            opp = _clean(ws.cell(row=r, column=7).value)
            if not host or not opp:
                continue
            matches.append(
                {
                    "host": host,
                    "opponent": opp,
                    "date": _jsonify_scalar(_clean(ws.cell(row=r, column=8).value)),
                    "venue": _clean(ws.cell(row=r, column=9).value),
                    "winner": _clean(ws.cell(row=r, column=10).value),
                    "motm": _clean(ws.cell(row=r, column=11).value),
                }
            )
    return {
        "name": "Emerging Talent League 2026",
        "squads": squads,
        "matches": matches,
        "champion": info.get("champion"),
        "runnerUp": info.get("runnerUp"),
        "status": "Completed" if info.get("champion") else "Ongoing",
    }


def _read_lone_warrior_matches(ws):
    """The bracket (Group A/B round robin, Quarter Finals, Semifinals, Grand
    Final) lives in 4 side-by-side columns, each schedule cell 2 columns to
    the left of its "Winner" cell ("Ruwanga vs Shamroz" in column A, winner
    in column C; same layout repeats at columns D/F, G/I, J/L). A row is a
    match only if its schedule cell contains " vs " - the date-only rows
    ("21st May") and the free-text scheduling note at the bottom aren't
    matches. The stage label above each column block ("Group A Matches",
    "Quarter Finals", "Semifinals", ...) gets reused for every match row
    beneath it until the next stage header in that same column."""
    matches = []
    stage_by_col = {1: None, 4: None, 7: None, 10: None}
    for r in range(1, ws.max_row + 1):
        for col in (1, 4, 7, 10):
            cell_val = _clean(ws.cell(row=r, column=col).value)
            if not isinstance(cell_val, str):
                continue
            if " vs " in cell_val.lower():
                winner = _clean(ws.cell(row=r, column=col + 2).value)
                matches.append({"stage": stage_by_col[col], "schedule": cell_val, "winner": winner})
            elif re.search(r"(matches|final)", cell_val, re.I):
                stage_by_col[col] = cell_val.strip()
    return matches


def get_lone_warrior(wb, tournament_updates):
    name = "IOCF Lone Warrior"
    info = tournament_updates.get("sections", {}).get("LONE WARRIOR 2026", {})
    representatives = {}
    matches = []
    if name in wb.sheetnames:
        ws = wb[name]
        for c in range(1, ws.max_column + 1):
            board = _clean(ws.cell(row=2, column=c).value)
            rep = _clean(ws.cell(row=3, column=c).value)
            if board and rep:
                if isinstance(board, str):
                    board = _BOARD_NAME_ALIASES.get(board.strip().lower(), board)
                representatives[board] = rep
        matches = _read_lone_warrior_matches(ws)
    completed_matches = [m for m in matches if m.get("winner")]
    return {
        "name": "IOCF Lone Warrior Season One",
        "representatives": representatives,
        "matches": matches,
        "totalMatches": len(completed_matches),
        "champion": info.get("champion"),
        "champion_board": info.get("champion_board"),
        "runnerUp": info.get("runnerUp"),
        "runnerUp_board": info.get("runnerUp_board"),
        "status": "Completed" if info.get("champion") else "Ongoing",
    }


# ---------------------------------------------------------------------------
# Top-level aggregate
# ---------------------------------------------------------------------------

def build_dashboard(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    boards, dismantled, credits_data = get_boards(wb)
    t20wc = get_t20_world_cup(wb)
    fixtures = get_fixtures(wb)
    franchise_leagues = get_franchise_leagues(wb)
    hall_of_fame = get_hall_of_fame(wb)
    continental_cups = get_continental_cups(wb)
    emerging = get_emerging_talent_league(wb, credits_data["tournamentUpdates"])
    lone_warrior = get_lone_warrior(wb, credits_data["tournamentUpdates"])
    board_rankings = get_board_rankings(wb)
    umpire_rankings = get_umpire_rankings(boards)
    head_to_head = get_head_to_head(wb)
    records = get_records(wb)

    # ---- tournaments (unified list for Tournaments module / featured cards)
    tournaments = []
    if t20wc:
        tournaments.append(
            {
                "id": "t20-world-cup-2026",
                "name": "T20 World Cup 2026",
                "category": "International",
                "season": "2026",
                "status": t20wc["status"],
                "champion": t20wc["champion"],
                "runnerUp": t20wc["runnerUp"],
                "totalMatches": t20wc["totalMatches"],
            }
        )
    tournaments.append(
        {
            "id": "emerging-talent-league-2026",
            "name": "Emerging Talent League 2026",
            "category": "League",
            "season": "2026",
            "status": emerging["status"],
            "champion": emerging["champion"],
            "runnerUp": emerging["runnerUp"],
            "totalMatches": len([m for m in emerging["matches"] if m.get("winner")]),
        }
    )
    tournaments.append(
        {
            "id": "lone-warrior-2026",
            "name": "IOCF Lone Warrior Season One",
            "category": "Knockout",
            "season": "2026",
            "status": lone_warrior["status"],
            "champion": lone_warrior["champion"],
            "runnerUp": lone_warrior["runnerUp"],
            "totalMatches": lone_warrior["totalMatches"],
        }
    )
    for cup in continental_cups:
        tournaments.append(
            {
                "id": cup["id"],
                "name": cup["name"],
                "category": "Continental Cup",
                "season": cup["season"],
                "status": cup["status"],
                "champion": cup["champion"],
                "runnerUp": None,
                "totalMatches": cup["totalMatches"],
            }
        )
    for lg in franchise_leagues:
        tournaments.append(
            {
                "id": lg["id"],
                "name": lg["name"],
                "category": "Franchise League",
                "season": lg["season"],
                "status": lg["status"],
                "champion": lg["champion"],
                "runnerUp": lg["runnerUp"],
                "totalMatches": lg["totalMatches"],
            }
        )

    # ---- totals
    total_stadiums = len({s.split("|")[0].strip() for b in boards for s in b["stadiums"]})
    completed_wtc_matches = [
        r for r in fixtures["tests"].get("worldTestChampionshipCompletedMatches", []) if r.get("Winners")
    ]
    total_matches = (
        (t20wc["totalMatches"] if t20wc else 0)
        + sum(_series_match_count(s.get("Series Result")) for s in fixtures["series"].get("completedSeries", []))
        + len(completed_wtc_matches)
        + sum(len([m for m in lg["matches"] if m.get("Winner")]) for lg in franchise_leagues)
        + len([m for m in emerging["matches"] if m.get("winner")])
        + lone_warrior["totalMatches"]
        + sum(len([m for m in cup["matches"] if m.get("Winner")]) for cup in continental_cups)
    )
    total_championships = len([t for t in tournaments if t.get("champion")])
    total_credits = sum(b["credits"] or 0 for b in boards)

    stats = {
        "totalBoards": len(boards),
        "totalPlayers": sum(b["playersCount"] for b in boards),
        "totalStadiums": total_stadiums,
        "totalTournaments": len(tournaments),
        "totalCredits": total_credits,
        "totalMatches": total_matches,
        "totalChampionships": total_championships,
    }

    # ---- latest updates / news feed (derived, not fabricated)
    news = []
    for t in tournaments:
        if t.get("champion"):
            news.append(
                {
                    "type": "tournament",
                    "headline": f"{t['champion']} — champions of {t['name']}",
                    "detail": f"Runners-up: {t.get('runnerUp')}" if t.get("runnerUp") else None,
                }
            )
    for b in boards:
        for tr in b["transfers"][:2]:
            news.append({"type": "transfer", "headline": f"{b['name']}: {tr}", "detail": None})
    for row in fixtures["series"].get("completedSeries", [])[-6:]:
        if row.get("Series Name") and row.get("Winners"):
            news.append(
                {
                    "type": "series",
                    "headline": f"{row.get('Series Name')}: {row.get('Winners')} beat {row.get('Runners', row.get('Opponents',''))}",
                    "detail": row.get("Dates"),
                }
            )
    for row in fixtures["tests"].get("worldTestChampionshipCompletedMatches", [])[-6:]:
        if row.get("Test Name") and row.get("Winners"):
            news.append(
                {
                    "type": "test",
                    "headline": f"{row.get('Test Name')}: {row.get('Winners')} won",
                    "detail": row.get("Dates"),
                }
            )

    # ---- upcoming matches (series + tests), simplified & merged
    upcoming = []
    for row in fixtures["series"].get("upcomingSeries", []):
        upcoming.append(
            {
                "type": "series",
                "name": row.get("Series Name"),
                "host": row.get("Hosting Board"),
                "opponents": row.get("Opponents"),
                "dates": row.get("Dates"),
                "format": row.get("Format"),
            }
        )
    for row in fixtures["tests"].get("upcomingIocfWorldTestChampionshipMatches", []):
        upcoming.append(
            {
                "type": "test",
                "name": row.get("Test Name"),
                "host": row.get("Hosting Board"),
                "opponents": row.get("Opponents"),
                "dates": row.get("Dates"),
                "format": "Test",
            }
        )

    return {
        "generatedAt": None,  # filled by server.py from file mtime
        "stats": stats,
        "boards": boards,
        "dismantledBoards": dismantled,
        "rankings": credits_data["rankings"],
        "boardRankings": board_rankings,
        "umpireRankings": umpire_rankings,
        "headToHead": head_to_head,
        "records": records,
        "tournaments": tournaments,
        "t20WorldCup": t20wc,
        "franchiseLeagues": franchise_leagues,
        "continentalCups": continental_cups,
        "emergingTalentLeague": emerging,
        "loneWarrior": lone_warrior,
        "fixtures": fixtures,
        "upcomingMatches": upcoming,
        "hallOfFame": hall_of_fame,
        "news": news,
    }


if __name__ == "__main__":
    import json
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else "IOCF_ALL_BOARDS.xlsx"
    data = build_dashboard(path)
    print(json.dumps(data, indent=2, default=str)[:4000])
